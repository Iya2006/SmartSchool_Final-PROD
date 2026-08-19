"""
SMARTSCHOOL — Lecture d'un fichier tabulaire téléversé (CSV / Excel).

Les résultats d'examen national arrivent de la Direction Préfectorale sous la
forme d'un tableau : parfois un vrai .xlsx, le plus souvent un export CSV
d'Excel francophone — séparé par des points-virgules et encodé en Windows-1252,
pas en UTF-8. Une lecture naïve `contenu.decode("utf-8").split(",")` échoue sur
les deux points à la fois, et l'école n'a aucun moyen de comprendre pourquoi.

Ce module ne fait qu'une chose : rendre un fichier téléversé sous forme de
lignes normalisées `{en-tête minuscule sans accent: valeur}`. Il ne connaît
aucune règle métier — l'appelant décide de ce que valent les colonnes.
"""
import csv
import io
import re
from typing import Dict, List, Tuple

_ACCENTS = "àáâãäåçèéêëìíîïñòóôõöùúûüýÿ"
_SANS = "aaaaaaceeeeiiiinooooouuuuyy"
_TABLE_ACCENTS = str.maketrans(_ACCENTS, _SANS)

# Encodages tentés dans l'ordre : un CSV exporté par Excel en français est en
# cp1252 neuf fois sur dix, mais un fichier passé par Google Sheets est en UTF-8.
_ENCODAGES = ("utf-8-sig", "utf-8", "cp1252", "latin-1")

# Séparateurs tentés : Excel francophone écrit `;`, l'export standard `,`,
# un copier-coller depuis un tableur donne une tabulation.
_SEPARATEURS = (";", ",", "\t")


class FichierIllisible(Exception):
    """Le fichier n'a pas pu être lu — le message est destiné à l'utilisateur."""


def normaliser_entete(texte: str) -> str:
    """`"N° Matricule"` -> `"n matricule"`, `"Résultat  final"` -> `"resultat final"`.

    Casse, accents, espaces multiples ET ponctuation sont neutralisés : les
    en-têtes réellement rencontrés s'écrivent `N°`, `N.`, `Nº` ou `No` selon qui
    a tapé le fichier, et une colonne manquée pour un degré Unicode fait passer
    tout un import pour un échec de rapprochement.

    Appliqué aussi aux matricules comparés de part et d'autre : `ELV-00013` et
    `ELV 00013` désignent le même élève.
    """
    brut = str(texte or "").lower().translate(_TABLE_ACCENTS)
    propre = "".join(c if (c.isalnum() or c.isspace()) else " " for c in brut)
    return " ".join(propre.split())


# Suffixes ordinaux français collés à un nombre (1er, 1ère, 2ème, 8e, 2nde…).
# Après neutralisation des accents ils s'écrivent « ere, eme, er, e, nde, nd… ».
# Ordonnés du plus long au plus court pour que « ere » l'emporte sur « e ».
_ORDINAL = re.compile(r"(\d+)(?:ere|eme|iere|ieme|nde|er|nd|es|e)(?=\s|$)")


def normaliser_classe(texte: str) -> str:
    """Comme `normaliser_entete`, mais insensible aux ORDINAUX collés aux nombres.

    « 1ère Année », « 1 ANNEE », « 1ere annee » désignent la même classe ; de
    même « 8ème Année » et « 8E ANNEE ». On réduit le nombre+suffixe au seul
    nombre : « 1ere annee » -> « 1 annee », « 8e annee » -> « 8 annee ». Sans ça,
    un fichier écrit « 1 ANNEE » ne retrouvait pas la classe « 1ère Année ».
    """
    base = normaliser_entete(texte)
    return " ".join(_ORDINAL.sub(r"\1 ", base).split())


def _decoder(contenu: bytes) -> str:
    for enc in _ENCODAGES:
        try:
            return contenu.decode(enc)
        except UnicodeDecodeError:
            continue
    raise FichierIllisible(
        "Encodage du fichier non reconnu. Enregistrez-le en CSV UTF-8 ou en .xlsx."
    )


def _lire_csv(contenu: bytes) -> List[List[str]]:
    texte = _decoder(contenu)
    premiere = next((l for l in texte.splitlines() if l.strip()), "")
    # On choisit le séparateur qui découpe réellement la première ligne : le
    # Sniffer de csv se trompe régulièrement sur un fichier à une seule colonne.
    separateur = max(_SEPARATEURS, key=premiere.count)
    if premiere.count(separateur) == 0:
        separateur = ";"
    return [ligne for ligne in csv.reader(io.StringIO(texte), delimiter=separateur)]


def _lire_xlsx(contenu: bytes) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise FichierIllisible(
            "Lecture des fichiers Excel indisponible sur ce serveur. "
            "Enregistrez le fichier au format CSV et réessayez."
        )
    classeur = load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    feuille = classeur.active
    return [
        ["" if c is None else str(c) for c in ligne]
        for ligne in feuille.iter_rows(values_only=True)
    ]


def lire_lignes(nom_fichier: str, contenu: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    """Retourne `(colonnes, lignes)` où chaque ligne est un dict indexé par
    en-tête normalisé. Les lignes entièrement vides sont ignorées.
    """
    if not contenu:
        raise FichierIllisible("Fichier vide.")
    nom = (nom_fichier or "").lower()
    if nom.endswith((".xlsx", ".xlsm")):
        brut = _lire_xlsx(contenu)
    elif nom.endswith(".xls"):
        raise FichierIllisible(
            "Format .xls (Excel 97) non pris en charge. "
            "Enregistrez le fichier en .xlsx ou en CSV."
        )
    else:
        brut = _lire_csv(contenu)

    brut = [l for l in brut if any(str(c).strip() for c in l)]
    if not brut:
        raise FichierIllisible("Aucune ligne exploitable dans le fichier.")

    colonnes = [normaliser_entete(c) for c in brut[0]]
    lignes = []
    for valeurs in brut[1:]:
        ligne = {
            colonnes[i]: str(valeurs[i]).strip()
            for i in range(min(len(colonnes), len(valeurs)))
            if colonnes[i]
        }
        if any(ligne.values()):
            lignes.append(ligne)
    return colonnes, lignes


def valeur(ligne: Dict[str, str], *noms_possibles: str) -> str:
    """Première valeur non vide parmi plusieurs noms de colonne acceptés.

    Les fichiers reçus n'ont jamais deux fois le même intitulé d'une année sur
    l'autre ("Matricule", "N° Matricule", "Numéro"...) : plutôt qu'imposer un
    gabarit strict, on accepte les variantes courantes.
    """
    for nom in noms_possibles:
        v = ligne.get(normaliser_entete(nom), "")
        if v:
            return v
    return ""
